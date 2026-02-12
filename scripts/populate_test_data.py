#!/usr/bin/env python3
"""
Populate workbook.xlsx with test data.
Account numbers are random 7-digit integers.
"""
import random
import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta

def generate_account_numbers(count=20):
    """Generate random 7-digit account numbers"""
    return [random.randint(1000000, 9999999) for _ in range(count)]

def populate_test_data():
    # Remove old workbook if it exists
    if os.path.exists("workbook.xlsx"):
        os.remove("workbook.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Generate random account numbers
    accounts = generate_account_numbers(20)
    staff_names = ["Alice Johnson", "Bob Smith", "Carol Williams", "David Brown", "Emma Davis"]
    departments = ["Accounting", "Operations", "Finance", "Admin"]
    allocation_types = ["Direct", "Indirect", "Overhead", "Adjustment"]
    note_types = ["Comment", "Alert", "Reconciliation", "Internal"]
    
    # 1. tblAccounts_Current
    ws.append(["Account_Number", "Account_Name", "Current_Balance", "Department"])
    for acc_num in accounts:
        ws.append([
            acc_num,
            f"Account_{acc_num}",
            round(random.uniform(1000, 100000), 2),
            random.choice(departments)
        ])
    tab = Table(displayName="tblAccounts_Current", ref=f"A1:D{len(accounts)+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # 2. tblStaff
    ws = wb.create_sheet("Sheet2")
    ws.append(["Staff_ID", "Name", "Department", "Email"])
    for idx, name in enumerate(staff_names, 1):
        ws.append([idx, name, random.choice(departments), f"{name.lower().replace(' ', '.')}@company.com"])
    tab = Table(displayName="tblStaff", ref=f"A1:D{len(staff_names)+1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    
    # 3. tblAllocations
    ws = wb.create_sheet("Sheet3")
    ws.append(["Allocation_ID", "Account_Number", "Amount", "Type", "Date_Allocated"])
    for idx in range(1, 31):
        base_date = datetime(2026, 1, 1)
        alloc_date = base_date + timedelta(days=random.randint(0, 41))
        ws.append([
            idx,
            random.choice(accounts),
            round(random.uniform(100, 10000), 2),
            random.choice(allocation_types),
            alloc_date.strftime("%Y-%m-%d")
        ])
    tab = Table(displayName="tblAllocations", ref=f"A1:E{30+1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    
    # 4. tblNotes
    ws = wb.create_sheet("Sheet4")
    ws.append(["Note_ID", "Account_Number", "Note_Type", "Note_Text", "Created_By", "Created_Date"])
    for idx in range(1, 21):
        base_date = datetime(2026, 1, 1)
        note_date = base_date + timedelta(days=random.randint(0, 41))
        ws.append([
            idx,
            random.choice(accounts),
            random.choice(note_types),
            f"Test note {idx}",
            random.choice(staff_names),
            note_date.strftime("%Y-%m-%d")
        ])
    tab = Table(displayName="tblNotes", ref=f"A1:F{20+1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    
    # 5. tblNoteTypes
    ws = wb.create_sheet("Sheet5")
    ws.append(["Type_ID", "Type_Name", "Description"])
    for idx, note_type in enumerate(note_types, 1):
        ws.append([idx, note_type, f"Type for {note_type} notes"])
    tab = Table(displayName="tblNoteTypes", ref=f"A1:C{len(note_types)+1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    
    # 6. tblConfig
    ws = wb.create_sheet("Sheet6")
    ws.append(["Setting_Name", "Setting_Value", "Description"])
    config_data = [
        ("Fiscal_Year_Start", "2026-01-01", "Start of fiscal year"),
        ("Fiscal_Year_End", "2026-12-31", "End of fiscal year"),
        ("Reconciliation_Frequency", "Monthly", "How often reconciliation runs"),
        ("Default_Currency", "USD", "Currency for all amounts"),
    ]
    for setting, value, desc in config_data:
        ws.append([setting, value, desc])
    tab = Table(displayName="tblConfig", ref=f"A1:C{len(config_data)+1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    
    # 7. tblRanges_Default
    ws = wb.create_sheet("Sheet7")
    ws.append(["Range_ID", "Range_Name", "Min_Value", "Max_Value"])
    ranges = [
        ("Small", 0, 1000),
        ("Medium", 1001, 10000),
        ("Large", 10001, 100000),
        ("Premium", 100001, 999999),
    ]
    for idx, (name, min_val, max_val) in enumerate(ranges, 1):
        ws.append([idx, name, min_val, max_val])
    tab = Table(displayName="tblRanges_Default", ref=f"A1:D{len(ranges)+1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    
    # 8. tblReconIssues
    ws = wb.create_sheet("Sheet8")
    ws.append(["Issue_ID", "Account_Number", "Issue_Description", "Severity", "Date_Found", "Status"])
    for idx in range(1, 6):
        base_date = datetime(2026, 1, 1)
        issue_date = base_date + timedelta(days=random.randint(0, 41))
        ws.append([
            idx,
            random.choice(accounts),
            f"Reconciliation issue {idx}",
            random.choice(["Low", "Medium", "High"]),
            issue_date.strftime("%Y-%m-%d"),
            "Open"
        ])
    tab = Table(displayName="tblReconIssues", ref=f"A1:F{5+1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    
    # 9. tblNoteSyncQueue
    ws = wb.create_sheet("Sheet9")
    ws.append(["Queue_ID", "Note_ID", "Operation", "Status", "Queued_Date"])
    ws.append(["", "", "", "", ""])  # Empty row to make table valid
    tab = Table(displayName="tblNoteSyncQueue", ref="A1:E2")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    
    # 10. tblNoteSyncLog
    ws = wb.create_sheet("Sheet10")
    ws.append(["Log_ID", "Note_ID", "Operation", "Result", "Sync_Date"])
    ws.append(["", "", "", "", ""])  # Empty row to make table valid
    tab = Table(displayName="tblNoteSyncLog", ref="A1:E2")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    
    # 11. tblRangeOverrides
    ws = wb.create_sheet("Sheet11")
    ws.append(["Override_ID", "Account_Number", "Range_Override", "Reason", "Effective_Date"])
    ws.append(["", "", "", "", ""])  # Empty row to make table valid
    tab = Table(displayName="tblRangeOverrides", ref="A1:E2")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    
    # 12. tblDigitMap
    ws = wb.create_sheet("Sheet12")
    ws.append(["Map_ID", "Digit_Position", "Meaning", "Valid_Values"])
    digit_mappings = [
        (1, 1, "Company Code", "1-9"),
        (2, 2, "Department", "0-9"),
        (3, 3, "SubDept", "0-9"),
        (4, 4, "Cost Center", "0-9"),
        (5, 5, "Account Type", "0-9"),
        (6, 6, "Reserved", "0-9"),
        (7, 7, "Sequence", "0-9"),
    ]
    for map_id, pos, meaning, values in digit_mappings:
        ws.append([map_id, pos, meaning, values])
    tab = Table(displayName="tblDigitMap", ref=f"A1:D{len(digit_mappings)+1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    
    wb.save("workbook.xlsx")
    print("✓ Test data populated successfully!")
    print(f"  - Generated {len(accounts)} unique account numbers (7-digit random)")
    print(f"  - Created 12 tables with sample data")
    print("  - Workbook saved to workbook.xlsx")

if __name__ == "__main__":
    populate_test_data()
